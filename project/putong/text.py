import re
import requests
import time
from selenium import webdriver
from openpyxl import Workbook
from openpyxl import load_workbook
class SinaSpider():
    def start_requests(self):
        url_list=[]
        for i in range(1,3):
            start_url = 'http://vip.stock.finance.sina.com.cn/quotes_service/api/json_v2.php/Market_Center.getHQNodeData?page={}&num=40&sort=changepercent&asc=0&node=hs_a&symbol=&_s_r_a=init'.format(i)
            url_list.append(start_url)
        return url_list

    def parse(self, url_list):
        symbols_list =[]
        for url in url_list:
            reponse = requests.get(url)
            text=reponse.text

            symbol_list = re.findall('symbol:"(.*?)"',text)
            symbols_list.append(symbol_list)
        return symbols_list
        # for symbol in symbol_list:
        #     url = 'http://finance.sina.com.cn/realstock/company/%s/nc.shtml'%symbol
        #     print(url)
    def page(self,symbols_list):
        wb = load_workbook('e:\\20181017.xlsx')
        ws = wb.active
        ws.append(['名称和代码', '主力买入', '主力卖出', '散户买入', '散户卖出', '主力买入比例', '主力卖出比例', '散户买入比例', '散户卖出比例', '散单','小单','大单','特大单',
                   '流通盘比例（散单）', '流通盘比例（小单）', '流通盘比例（大单）', '流通盘比例（特大单）', '换手率比例（散单）', '换手率比例（小单）', '换手率比例（大单）', '换手率比例（特大单）'])

        for list in symbols_list:
            z=0
            for k in list:
                z +=1
                print("插入第 %d 条数据"%z)
                chrome_options = webdriver.ChromeOptions()
                # 使用headless无界面浏览器模式
                chrome_options.add_argument('--headless')
                chrome_options.add_argument('--disable-gpu')


                driver = webdriver.Chrome(chrome_options=chrome_options)
                url = 'http://finance.sina.com.cn/realstock/company/{}/nc.shtml'.format(k)
                print(url)
                try:
                    driver.get(url)

                #time.sleep(3)
                    a = driver.find_elements_by_xpath('//div[@id="MRFlow"]//tbody/tr[2]/td')
                    b = driver.find_elements_by_xpath('//div[@id="MRFlow"]//tbody/tr[3]/td')
                    c = driver.find_elements_by_xpath('//div[@id="FLFlow"]//tbody/tr[2]/td')
                    d = driver.find_elements_by_xpath('//div[@id="FLFlow"]//tbody/tr[3]/td')
                    e = driver.find_elements_by_xpath('//div[@id="FLFlow"]//tbody/tr[4]/td')
                    g = driver.find_elements_by_xpath('//h1[@id="stockName"]')
                    f = g + a + b + c + d + e
                    #wb = Workbook()


                    data = []

                    for n in f:
                        data.append(n.text)

                    print(data)
                    time.sleep(3)
                    driver.quit()

                    ws.append(data)
                    wb.save('e:\\20181017.xlsx')
                except:
                    print("出错啦")


            # for k in range(5):
            #     wb = Workbook()
            #
            #     ws = wb.active
            #     ws.append(data)
            # wb.save('e:\\20181014.xlsx')
            # return data
            # time.sleep(3)
            # driver.quit()


    # def save(self,m):
    #     # wb = Workbook()
    #     # ws = wb.active
    #     print(m)
        # if data is not None:
        #
        #     # ws.append(
        #     #     ['代码', '名称', '最新价', '今日涨跌幅', '净额(主力)', '净占比(主力)', '净额(超大单)', '净占比(超大单)', '净额(大单)', '净占比(大单)', '净额(中单)',
        #     #      '净占比(中单)', '净额(小单)', '净占比(小单)', '时间'])
        #     line = data
        #     print("开始插入")
        #     ws.append(line)  # 将数据以行的形式添加到xlsx中
        # else:
        #     wb.save('e:\\20181014.xlsx')
    def run(self):
        url_list = self.start_requests()
        symbols_list=self.parse(url_list)
        self.page(symbols_list)

if __name__ == '__main__':
    spider = SinaSpider()
    spider.run()